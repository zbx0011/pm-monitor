"""
使用 CME 价格 Excel 数据更新合约收敛分析图表
只保留国内外数据都有的日期（交集）
"""

import json
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def load_cme_data(excel_path):
    """从 Excel 加载 CME 价格数据"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # 读取数据
    data = {}
    for row in ws.iter_rows(min_row=6, max_col=7):  # 跳过头部
        date = row[0].value
        if date is None or not isinstance(date, datetime):
            continue
        
        date_str = date.strftime('%Y-%m-%d')
        data[date_str] = {
            'GCZ24': row[1].value,  # 黄金2412 -> GCZ24
            'GCM24': row[2].value,  # 黄金2406 -> GCM24
            'GCZ23': row[3].value,  # 黄金2312 -> GCZ23
            'GCM23': row[4].value,  # 黄金2306 -> GCM23
            'SIZ24': row[5].value,  # 白银2412 -> SIZ24
            'SIM24': row[6].value,  # 白银2406 -> SIM24
        }
    
    return data

def load_existing_convergence_data(json_path):
    """加载现有的收敛数据"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def update_convergence_data(cme_data, existing_data, exchange_rate=7.04):
    """更新收敛数据，只保留国内外数据都有的日期"""
    
    # 合约映射：国内合约 -> CME合约代码
    contract_mapping = {
        'AU2412': 'GCZ24',  # 黄金 2024年12月 -> COMEX Gold Dec 2024
        'AU2406': 'GCM24',  # 黄金 2024年6月 -> COMEX Gold Jun 2024  
        'AU2312': 'GCZ23',  # 黄金 2023年12月 -> COMEX Gold Dec 2023
        'AU2306': 'GCM23',  # 黄金 2023年6月 -> COMEX Gold Jun 2023
        'AG2412': 'SIZ24',  # 白银 2024年12月 -> COMEX Silver Dec 2024
        'AG2406': 'SIM24',  # 白银 2024年6月 -> COMEX Silver Jun 2024
    }
    
    # 克/盎司转换
    OZ_TO_GRAM = 31.1035
    
    for contract in existing_data['contracts']:
        symbol = contract['symbol']
        cme_code = contract_mapping.get(symbol)
        
        if not cme_code:
            print(f"跳过未知合约: {symbol}")
            continue
            
        print(f"\n处理合约: {symbol} -> {cme_code}")
        
        # 判断是黄金还是白银
        is_gold = symbol.startswith('AU')
        
        # 筛选数据：只保留国内外都有数据的日期
        new_history = []
        matched_count = 0
        
        for point in contract['history']:
            date = point['date']
            close = point['close']  # 国内价格
            
            # 检查CME是否有这一天的数据
            if date in cme_data:
                cme_price = cme_data[date].get(cme_code)
                if cme_price is not None and cme_price > 0 and close is not None and close > 0:
                    # 计算国际价格
                    if is_gold:
                        # 黄金：国内是元/克，CME是美元/盎司
                        # CME 美元/oz -> 人民币/克
                        intl_cny = (cme_price * exchange_rate) / OZ_TO_GRAM
                    else:
                        # 白银：国内是元/千克，CME是美元/盎司
                        # CME 美元/oz -> 人民币/千克
                        intl_cny = (cme_price * exchange_rate) / OZ_TO_GRAM * 1000
                    
                    # 计算溢价率
                    spread_pct = ((close - intl_cny) / intl_cny) * 100
                    
                    new_history.append({
                        'date': date,
                        'close': close,
                        'intl_cny': intl_cny,
                        'spread_pct': spread_pct
                    })
                    matched_count += 1
        
        print(f"  原始数据点: {len(contract['history'])}")
        print(f"  匹配数据点: {matched_count}")
        
        # 更新历史数据
        if new_history:
            contract['history'] = new_history
            contract['start_date'] = new_history[0]['date']
            contract['end_date'] = new_history[-1]['date']
            contract['days'] = len(new_history)
            contract['start_spread'] = new_history[0]['spread_pct']
            contract['end_spread'] = new_history[-1]['spread_pct']
            contract['convergence'] = contract['start_spread'] - contract['end_spread']
    
    # 更新时间戳
    existing_data['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    return existing_data

def main():
    excel_path = r'E:\微信\文件\WeChat Files\q177312837\FileStorage\File\2025-12\价格导出.xlsx'
    json_path = r'e:\项目\币圈等监控系统\contract_convergence_data.json'
    
    print("加载 CME 数据...")
    cme_data = load_cme_data(excel_path)
    print(f"加载了 {len(cme_data)} 条 CME 价格记录")
    
    # 检查每个合约的有效数据量
    print("\n各合约有效数据量:")
    for code in ['GCZ24', 'GCM24', 'GCZ23', 'GCM23', 'SIZ24', 'SIM24']:
        valid_count = sum(1 for date, row in cme_data.items() if row.get(code) is not None and row.get(code) > 0)
        print(f"  {code}: {valid_count} 条")
    
    print("\n加载现有收敛数据...")
    existing_data = load_existing_convergence_data(json_path)
    print(f"现有 {len(existing_data['contracts'])} 个合约")
    
    print("\n更新收敛数据（只保留交集）...")
    updated_data = update_convergence_data(cme_data, existing_data)
    
    # 保存更新后的数据
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(updated_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n已保存到: {json_path}")
    
    # 打印汇总
    print("\n=== 更新汇总 ===")
    for contract in updated_data['contracts']:
        print(f"{contract['symbol']}: {contract['days']}天, 溢价 {contract['start_spread']:.2f}% → {contract['end_spread']:.2f}%")

if __name__ == '__main__':
    main()
